from openpyxl import Workbook

org=input("please setting your organization name: ")
prj=input("please setting your repository/project name: ")

wb = Workbook()

ws_reg=wb.create_sheet("Register")
ws_mem=wb.create_sheet("Memory")
ws_afio=wb.create_sheet("AFIO")
ws_fpga=wb.create_sheet("FPGA")

ws_fpga['A1']="I/O Port"
ws_fpga['B1']="Pin Number"
ws_fpga['C1']="Site"
ws_fpga['D1']="Site Type"
ws_fpga['E1']="Min/Max Trace Delay (ps)"
ws_fpga['F1']="Trace Length (um)"
ws_fpga['G1']="Prohibit"
ws_fpga['H1']="Interface"
ws_fpga['I1']="Signal Name"
ws_fpga['J1']="Direction"
ws_fpga['K1']="DiffPair Type"
ws_fpga['L1']="DiffPair Signal"
ws_fpga['M1']="IO Standard"
ws_fpga['N1']="Drive"
ws_fpga['O1']="Slew Rate"
ws_fpga['P1']="OUTPUT_IMPEDANCE"
ws_fpga['Q1']="PRE_EMPHASIS"
ws_fpga['R1']="LVDS_PRE_EMPHASIS"
ws_fpga['S1']="Pull Type"
ws_fpga['T1']="IN_TERM"
ws_fpga['U1']="DQS_BIAS"
ws_fpga['V1']="DIFF_TERM"
ws_fpga['W1']="OFFCHIP_TERM"
ws_fpga['X1']="Board Signal"
ws_fpga['Y1']="Board Voltage"
ws_fpga['Z1']="ODT"

wb.save('asic.xlsx')
